import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
import os
from datetime import datetime

def calcular_ickl(df):
    """
    Calcula el Índice Check List OCRA (ICKL) basado en los datos del DataFrame.
    
    :param df: DataFrame que contiene los datos necesarios para el cálculo
    :return: Valor del Índice Check List OCRA (ICKL)
    """
    # Obtener valores para el cálculo de ICKL
    fr = df['Factor de Recuperacion'].mean()  # Promedio de los valores si hay múltiples registros
    ff = df['FF'].mean()
    ffz = df['Factor de Fuerza'].mean()
    fp = df['Factor de Posturas y Movimientos'].mean()
    fc = df['Factor de Riesgos Adicionales'].mean()
    md = df['Multiplicador de Duracion'].mean()

    # Calcular el Índice Check List OCRA (ICKL)
    return (fr + ff + ffz + fp + fc) * md

def determinar_nivel_riesgo(ickl):
    """
    Determina el nivel de riesgo basado en el Índice Check List OCRA (ICKL).
    
    :param ickl: Índice Check List OCRA (ICKL)
    :return: Tuple con el nivel de riesgo, acción recomendada y el índice OCRA equivalente
    """
    if ickl <= 5:
        return 'Óptimo', 'No se requiere', '≤ 1.5'
    elif 5.1 <= ickl <= 7.5:
        return 'Aceptable', 'No se requiere', '1.6 - 2.2'
    elif 7.6 <= ickl <= 11:
        return 'Incierto', 'Se recomienda un nuevo análisis o mejora del puesto', '2.3 - 3.5'
    elif 11.1 <= ickl <= 14:
        return 'Inaceptable Leve', 'Se recomienda mejora del puesto, supervisión médica y entrenamiento', '3.6 - 4.5'
    elif 14.1 <= 22.5:
        return 'Inaceptable Medio', 'Se recomienda mejora del puesto, supervisión médica y entrenamiento', '4.6 - 9'
    else:
        return 'Inaceptable Alto', 'Se recomienda mejora del puesto, supervisión médica y entrenamiento', '> 9'

def main():
    # Definir la carpeta donde están los archivos JSON
    json_folder = 'C:/Tesis/TestErgo/archivos_json'
    json_path = get_most_recent_file(json_folder, '.json')

    # Leer el archivo JSON
    with open(json_path, 'r') as file:
        data_list = json.load(file)

    # Verificar si data_list es una lista y si no está vacía
    if not isinstance(data_list, list) or len(data_list) == 0:
        raise ValueError("El archivo JSON no contiene una lista válida o está vacío.")

    # Convertir a DataFrame de pandas
    df = pd.DataFrame(data_list)

    # Calcular el Índice Check List OCRA (ICKL)
    ickl = calcular_ickl(df)

    # Determinar Nivel de Riesgo y Acción Recomendada
    nivel_riesgo, accion_recomendada, indice_ocra_equivalente = determinar_nivel_riesgo(ickl)

    # Obtener el nombre del video y la fecha actual
    nombre_video = df['analisis_video'][0].split('\\')[-1].replace('.mp4', '')
    fecha_hora = datetime.now().strftime('%Y%m%d_%H%M%S')

    # Construir el nombre del archivo Excel
    archivo_excel = f'analisisOcra_{nombre_video}_{fecha_hora}.xlsx'

    # Definir la ruta de la carpeta 'resultados' dentro de 'static'
    static_folder = 'static'
    resultados_folder = os.path.join(static_folder, 'resultados')

    # Crear la carpeta 'resultados' si no existe
    if not os.path.exists(resultados_folder):
        os.makedirs(resultados_folder)

    # Definir la ruta del archivo Excel en la carpeta 'resultados'
    excel_path = os.path.join(resultados_folder, archivo_excel)

    # Crear un nuevo libro de trabajo
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen OCRA"

    # Escribir encabezados
    headers = ["Índice Check List OCRA", "Nivel de Riesgo", "Acción Recomendada", "Índice OCRA Equivalente"]
    ws.append(headers)

    # Escribir la fila de resultados
    resultados = [ickl, nivel_riesgo, accion_recomendada, indice_ocra_equivalente]
    ws.append(resultados)

    # Estilos para la hoja de cálculo
    header_fill = PatternFill(start_color='4C47EA', end_color='4C47EA', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    alignment = Alignment(horizontal='center', vertical='center')

    # Aplicar estilo a los encabezados
    for col_num in range(1, 5):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = alignment

    # Aplicar formato a las celdas de datos
    for cell in ws[2]:
        cell.alignment = alignment
        if isinstance(cell.value, (int, float)):
            # Ejemplo de formato condicional para resaltar valores altos
            if cell.value > 14:
                cell.fill = PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')
            elif cell.value > 7.5:
                cell.fill = PatternFill(start_color='FFCC99', end_color='FFCC99', fill_type='solid')

    # Ajustar el ancho de las columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Guardar el archivo Excel
    wb.save(excel_path)

    print(f"El Excel se ha creado y guardado en {excel_path}.")

if __name__ == "__main__":
    main()